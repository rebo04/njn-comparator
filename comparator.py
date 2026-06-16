"""
NJN PPAP Comparator — core logic (importable by app.py or run standalone).
"""

import difflib
import re
import shutil, os, sys, glob
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

# ── Cake / Pastel Color Palette ───────────────────────────────────────────────
FILL_CHANGED  = PatternFill("solid", fgColor="FFF3B0")  # lemon chiffon
FILL_ADDED    = PatternFill("solid", fgColor="B5EAD7")  # mint
FILL_REMOVED  = PatternFill("solid", fgColor="FFAAB5")  # strawberry
FILL_SYSTEMIC = PatternFill("solid", fgColor="D6EAF8")  # light sky blue — systemic banner
FILL_RENAME   = PatternFill("solid", fgColor="E5D4F0")  # light lavender — global value-rename

FONT_CHANGED  = Font(name="Century Gothic", size=11, color="5C4000")
FONT_ADDED    = Font(name="Century Gothic", size=11, bold=True, color="1A5C3C")
FONT_REMOVED  = Font(name="Century Gothic", size=11, bold=True, color="7B001A", strike=True)

AUTHOR = "Arturo Rebolledo"

FILL_LEGEND_TITLE = PatternFill("solid", fgColor="2E4057")  # dark navy
FILL_LEGEND_CHG   = PatternFill("solid", fgColor="FFF3B0")
FILL_LEGEND_ADD   = PatternFill("solid", fgColor="B5EAD7")
FILL_LEGEND_DEL   = PatternFill("solid", fgColor="FFAAB5")
FILL_LEGEND_SYS   = PatternFill("solid", fgColor="D6EAF8")
FILL_LEGEND_RENAME = PatternFill("solid", fgColor="E5D4F0")
FILL_LEGEND_SAME  = PatternFill("solid", fgColor="F5F5F7")


def fmt(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


# ── Fix #1 & #2: robust sheet loading + formula-cache fallback ───────────────
def _load_worksheet(path, data_only):
    """Load workbook and return (wb, ws), trying 'NJN' first then active sheet."""
    wb = openpyxl.load_workbook(path, data_only=data_only)
    if "NJN" in wb.sheetnames:
        return wb, wb["NJN"]
    ws = wb.active
    if ws is not None:
        return wb, ws
    raise ValueError(
        f"Sheet 'NJN' not found and no active sheet available. "
        f"Available sheets: {wb.sheetnames}"
    )


def read_values(path):
    """
    Read cell values from the NJN sheet (or active sheet).

    Uses data_only=True to get computed values. When a cell returns None
    under data_only but the same cell contains a formula in the raw workbook,
    the raw formula string is used as the display value so it is visible in
    the diff output.

    Fix #8: wraps everything in try/except and raises a descriptive ValueError.
    """
    try:
        _, ws_data = _load_worksheet(path, data_only=True)
        _, ws_raw  = _load_worksheet(path, data_only=False)

        str_vals, raw_vals = {}, {}
        for row in ws_data.iter_rows():
            for cell in row:
                key = (cell.row, cell.column)
                val = cell.value

                # Fix #2: formula cache miss → treat as empty for comparison.
                # Using the raw formula string causes false "changed" flags when
                # one file has a cache and the other does not.
                if val is None:
                    raw_cell = ws_raw.cell(cell.row, cell.column)
                    raw_val  = raw_cell.value
                    if isinstance(raw_val, str) and raw_val.startswith("="):
                        val = None  # unknown — compare as empty, not as formula text

                str_vals[key] = fmt(val)
                raw_vals[key] = val

        max_row = ws_data.max_row
        max_col = ws_data.max_column
        return str_vals, raw_vals, max_row, max_col

    except (ValueError, KeyError):
        raise   # re-raise our own descriptive errors as-is
    except Exception as exc:
        raise ValueError(
            f"Could not read '{os.path.basename(path)}': {exc}"
        ) from exc


# ── Fix #3: dynamic column detection ─────────────────────────────────────────
_PART_KEYWORDS   = {"PART", "PART NUMBER", "NUMERO", "NUM", "P/N", "PN"}
_LEVEL_KEYWORDS  = {"LEVEL", "NIVEL", "LVL"}
_PARENT_KEYWORDS = {"PARENT", "PADRE", "PARENT PN", "PADRE PN"}

def _detect_columns(vals, max_col, header_end):
    """
    Scan header rows (1..header_end) for column labels and return
    (col_pn, col_level, col_parent).  Falls back to (5, 3, 4) if not found.
    """
    col_pn, col_level, col_parent = 5, 3, 4   # hardcoded defaults

    for r in range(1, header_end + 1):
        for c in range(1, max_col + 1):
            raw = (vals.get((r, c)) or "").upper().strip()
            if raw in _PART_KEYWORDS and col_pn == 5:
                col_pn = c
            if raw in _LEVEL_KEYWORDS and col_level == 3:
                col_level = c
            if raw in _PARENT_KEYWORDS and col_parent == 4:
                col_parent = c

    return col_pn, col_level, col_parent


# ── Fix #4: dynamic header-boundary detection ─────────────────────────────────
def _detect_header_end(vals, max_row, col_pn, col_level, max_search=15):
    """
    Return the last header row index (rows up to and including this index are
    treated as fixed header rows).  Looks for the first row after row 1 that
    has a numeric-looking value in col_level OR a non-empty value in col_pn
    that looks like a part number (digits/alphanumeric, len > 3).
    Caps search at min(max_search, max_row).
    """
    limit = min(max_search, max_row)
    for r in range(2, limit + 1):
        level_val = (vals.get((r, col_level)) or "").strip()
        pn_val    = (vals.get((r, col_pn))    or "").strip()

        # Numeric level value → data row starts here
        if level_val and re.match(r'^\d+(\.\d+)*$', level_val):
            return r - 1

        # Part-number-like value (at least 4 alphanum chars, mostly digits)
        if pn_val and len(pn_val) >= 4 and sum(ch.isdigit() for ch in pn_val) >= 3:
            return r - 1

    return 8   # safe fallback


def add_legend(ws, start_row, label_a, label_b, max_col, systemic_info, rename_info=None):
    end_col = max_col + 1
    r = start_row + 2

    # Title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
    t = ws.cell(r, 1, f"  {label_a}  →  {label_b}")
    t.fill = FILL_LEGEND_TITLE
    t.font = Font(name="Century Gothic", size=13, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 28
    r += 1

    # Timestamp + author
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
    s = ws.cell(r, 1, f"  NJN Comparator  ·  {datetime.now().strftime('%Y-%m-%d  %H:%M')}  ·  Compared by: {AUTHOR}")
    s.fill = PatternFill("solid", fgColor="3D5A73")
    s.font = Font(name="Century Gothic", size=9, color="FFFFFF")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 16
    r += 1

    items = [
        (FILL_LEGEND_CHG,    "CHANGED",   "Cell value was modified — hover to see the previous value"),
        (FILL_LEGEND_ADD,    "ADDED",     "Row is new in this revision"),
        (FILL_LEGEND_DEL,    "DELETED",   "Row was removed — shown in its original position"),
        (FILL_LEGEND_SYS,    "SYSTEMIC",  "Column changed across most rows — see blue banner above the data"),
        (FILL_LEGEND_RENAME, "RENAMED",   "Same value substitution repeated across many rows — counted once, see purple banner above the data"),
        (FILL_LEGEND_SAME,   "UNCHANGED", "No difference"),
    ]
    for fill, lbl, desc in items:
        ws.row_dimensions[r].height = 22
        sw = ws.cell(r, 1, f"  {lbl}")
        sw.fill = fill
        sw.font = Font(name="Century Gothic", size=10, bold=True, color="333333")
        sw.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=end_col)
        dc = ws.cell(r, 2, f"  {desc}")
        dc.fill = FILL_LEGEND_SAME
        dc.font = Font(name="Century Gothic", size=10, color="444444")
        dc.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    if systemic_info:
        ws.row_dimensions[r].height = 16
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
        si = ws.cell(r, 1, "  Systemic changes: " + "  |  ".join(
            f"Col {c}: '{old}' → '{new}'" for c, old, new in systemic_info
        ))
        si.fill = FILL_SYSTEMIC
        si.font = Font(name="Century Gothic", size=9, italic=True, color="FFFFFF")
        si.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    if rename_info:
        ws.row_dimensions[r].height = 16
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
        ri = ws.cell(r, 1, "  Global value renames: " + "  |  ".join(
            f"'{old}' → '{new}' ({cnt} occurrences)" for old, new, cnt in rename_info
        ))
        ri.fill = FILL_RENAME
        ri.font = Font(name="Century Gothic", size=9, italic=True, color="333333")
        ri.alignment = Alignment(horizontal="left", vertical="center")


def _strip_revision_suffix(s):
    """
    Strip a trailing single-letter revision suffix from a part number, so that
    e.g. "656475500F" and "656475500G" (or "690BKTP-C" and "690BKTP-B") are
    recognized as the same underlying part at different revisions.

    Two patterns are recognized:
      1. Hyphen-separated suffix:  "690BKTP-C"   -> "690BKTP"
      2. Bare trailing letter:     "656475500F"  -> "656475500"
    Only fires on strings long enough that the stripped remainder is still a
    meaningful identifier (avoids mangling short codes).
    """
    if not s:
        return s
    m = re.match(r'^(.+)-[A-Za-z]$', s)
    if m and len(m.group(1)) >= 4:
        return m.group(1)
    if len(s) > 5 and s[-1].isupper() and s[-2:].isalnum():
        return s[:-1]
    return s


def _row_key(r, data, col_pn=5, col_level=3, col_parent=4, header_end=8):
    """
    Build a stable key for a data row.

    Fix #3: uses detected column positions instead of hardcoded 3/4/5.
    Fix #4: uses detected header_end instead of hardcoded 8.
    Fix #6: CONTENT fallback only uses first 3 non-empty values.
    Fix #9: strip trailing revision letter from BOTH parent and the row's own
            part number, so a part whose only change is its own revision
            suffix still aligns as "changed" instead of delete+insert.
    """
    if r <= header_end:
        return f"__HDR_{r:03d}"
    pn = (data.get(col_pn) or "").strip()
    if pn:
        level  = (data.get(col_level)  or "").strip()
        parent = (data.get(col_parent) or "").strip()
        parent = _strip_revision_suffix(parent)
        pn_key = _strip_revision_suffix(pn)
        return f"PART|{level}|{parent}|{pn_key}"
    label = (data.get(1) or "").strip()
    if label:
        return f"FOOTER|{label}"
    # Fix #6: use only first 3 non-empty values for stability
    non_empty = [str(v) for v in data.values() if v][:3]
    return "CONTENT|" + "|".join(non_empty)


def compare_and_export(path_a, path_b, out_path, label_a, label_b):
    """
    Compare two NJN Excel files and produce a highlighted output:
      - Lemon yellow cell  : value changed (old value shown in ghost row below)
      - Mint green row     : row added in path_b
      - Pink row           : row deleted from path_a, shown in-place
      - Sky-blue banner    : column that changed in ≥80% of rows (systemic)
      - Lavender cell/banner: the same (old_value, new_value) pair recurring
                              ≥5 times anywhere in the sheet (global rename,
                              e.g. a revision-letter bump propagating into
                              every row that references that part). Each
                              affected cell is still highlighted, but the
                              whole group counts as 1 toward n_changed.

    Returns (n_changed_cells, n_added_rows, n_removed_rows), where
    n_changed_cells already collapses global-rename groups to 1 each so the
    headline number reflects distinct signal rather than mechanical repeats.
    """
    vals_a, _,     max_row_a, max_col_a = read_values(path_a)
    vals_b, raw_b, max_row_b, max_col_b = read_values(path_b)
    max_col = max(max_col_a, max_col_b)

    # Fix #3 & #4: detect header boundary and column positions from file A
    # (use A as the reference since it's the older/base revision)
    # Quick pass: detect header end with default columns first, then refine
    col_pn_tmp, col_level_tmp, col_parent_tmp = 5, 3, 4
    header_end = _detect_header_end(vals_a, max_row_a, col_pn_tmp, col_level_tmp)
    col_pn, col_level, col_parent = _detect_columns(vals_a, max_col_a, header_end)
    # Re-detect header end now that we have the real column positions
    header_end = _detect_header_end(vals_a, max_row_a, col_pn, col_level)

    def get_rows(vals, max_row):
        out = []
        for r in range(1, max_row + 1):
            data = {c: vals.get((r, c), "") for c in range(1, max_col + 1)}
            if any(data.values()):
                out.append((r, data))
        return out

    rows_a = get_rows(vals_a, max_row_a)
    rows_b = get_rows(vals_b, max_row_b)

    keys_a = [_row_key(r, d, col_pn, col_level, col_parent, header_end) for r, d in rows_a]
    keys_b = [_row_key(r, d, col_pn, col_level, col_parent, header_end) for r, d in rows_b]

    # ── difflib alignment ────────────────────────────────────────────────────
    sm  = difflib.SequenceMatcher(None, keys_a, keys_b, autojunk=False)
    ops = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            for ia, ib in zip(range(i1, i2), range(j1, j2)):
                ops.append(("same",    rows_b[ib][0], rows_b[ib][1], rows_a[ia][1]))
        elif tag == "insert":
            for ib in range(j1, j2):
                ops.append(("added",   rows_b[ib][0], rows_b[ib][1], None))
        elif tag == "delete":
            for ia in range(i1, i2):
                ops.append(("deleted", None,           None,          rows_a[ia][1]))
        elif tag == "replace":
            # Fix #5: secondary SequenceMatcher for better inner-block alignment
            sub_a = [rows_a[ia] for ia in range(i1, i2)]
            sub_b = [rows_b[ib] for ib in range(j1, j2)]

            def _row_content_str(data):
                return " | ".join(str(v) for v in data.values() if v)

            keys_sub_a = [_row_content_str(d) for _, d in sub_a]
            keys_sub_b = [_row_content_str(d) for _, d in sub_b]

            inner_sm  = difflib.SequenceMatcher(None, keys_sub_a, keys_sub_b, autojunk=False)
            for itag, ii1, ii2, ij1, ij2 in inner_sm.get_opcodes():
                if itag == "equal":
                    for ka, kb in zip(range(ii1, ii2), range(ij1, ij2)):
                        ops.append(("same",    sub_b[kb][0], sub_b[kb][1], sub_a[ka][1]))
                elif itag == "insert":
                    for kb in range(ij1, ij2):
                        ops.append(("added",   sub_b[kb][0], sub_b[kb][1], None))
                elif itag == "delete":
                    for ka in range(ii1, ii2):
                        ops.append(("deleted", None,          None,         sub_a[ka][1]))
                elif itag == "replace":
                    pairs = min(ii2 - ii1, ij2 - ij1)
                    for k in range(pairs):
                        ops.append(("changed", sub_b[ij1+k][0], sub_b[ij1+k][1], sub_a[ii1+k][1]))
                    for ka in range(ii1 + pairs, ii2):
                        ops.append(("deleted", None, None, sub_a[ka][1]))
                    for kb in range(ij1 + pairs, ij2):
                        ops.append(("added",   sub_b[kb][0], sub_b[kb][1], None))

    # ── Systemic change detection ────────────────────────────────────────────
    # A column is "systemic" if it differs in ≥80% of matched row pairs AND
    # at least 5 rows are affected — likely a global revision-letter change.
    matched = [(da, db) for kind, _, db, da in ops
               if kind in ("same", "changed") and da is not None]
    systemic_cols: dict = {}   # col → (old_sample, new_sample)
    if len(matched) >= 5:
        for c in range(1, max_col + 1):
            diffs = [(da.get(c,""), db.get(c,"")) for da, db in matched
                     if da.get(c,"") != db.get(c,"")]
            if len(diffs) >= max(5, 0.8 * len(matched)):
                old_s = next((p[0] for p in diffs if p[0]), "")
                new_s = next((p[1] for p in diffs if p[1]), "")
                systemic_cols[c] = (old_s, new_s)

    # ── Global value-substitution detection ──────────────────────────────────
    # The column-based systemic detector above only fires when ONE column
    # changes in ≥80% of ALL rows. Real PPAP BOMs often have a column (e.g.
    # "parent part number") that holds many *different* values, so a single
    # value being renamed everywhere it occurs (e.g. a revision-letter bump
    # propagating into every row that references it) never reaches that
    # global 80% threshold — yet it's still conceptually ONE rename, just
    # repeated mechanically. Detect that here by grouping every per-cell
    # diff (across ALL rows/columns, not just systemic-column rows) by the
    # literal (old_value, new_value) pair. A pair recurring often enough is
    # treated as a single global rename: highlighted everywhere it appears,
    # but counted once in the headline number instead of once per occurrence.
    GLOBAL_RENAME_MIN_OCCURRENCES = 5
    rename_counts: dict = {}        # (old, new) -> occurrence count
    for da, db in matched:
        for c in range(1, max_col + 1):
            va, vb = da.get(c, ""), db.get(c, "")
            if va != vb and va and vb:
                rename_counts[(va, vb)] = rename_counts.get((va, vb), 0) + 1
    global_renames = {
        pair: cnt for pair, cnt in rename_counts.items()
        if cnt >= GLOBAL_RENAME_MIN_OCCURRENCES
    }

    # ── Build insertion map (deleted rows → before which B row) ──────────────
    insertions: dict = {}
    pending:    list = []
    for kind, r_b, _, data_a in ops:
        if kind == "deleted":
            pending.append(data_a)
        elif pending:
            insertions.setdefault(r_b, []).extend(pending)
            pending = []
    trailing_dels = pending

    # ── Copy newer file ───────────────────────────────────────────────────────
    shutil.copy(path_b, out_path)
    wb_out = openpyxl.load_workbook(out_path)

    # Fix #1: robust sheet selection for output workbook
    if "NJN" in wb_out.sheetnames:
        ws = wb_out["NJN"]
    elif wb_out.active is not None:
        ws = wb_out.active
    else:
        raise ValueError(
            f"Sheet 'NJN' not found in output workbook. "
            f"Available sheets: {wb_out.sheetnames}"
        )

    # Replace formula cells with their computed values (prevents "31-Dec-99" bug)
    for (r, c), raw_val in raw_b.items():
        cell = ws.cell(r, c)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = raw_val

    # ── Insert systemic/rename banner row(s) just below column-header row ─────
    # Use detected header_end instead of hardcoded 8
    banner_rows_needed = (1 if systemic_cols else 0) + (1 if global_renames else 0)
    if banner_rows_needed:
        banner_row = header_end + 1
        for _ in range(banner_rows_needed):
            ws.insert_rows(banner_row)

        cur_row = banner_row
        if systemic_cols:
            ws.row_dimensions[cur_row].height = 22
            desc = "  SYSTEMIC CHANGES (affect most rows):  " + "    ·    ".join(
                f"Col {c}  '{old}' → '{new}'"
                for c, (old, new) in systemic_cols.items()
            )
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row, end_column=max_col + 3)
            bc = ws.cell(cur_row, 1, desc)
            bc.fill = FILL_SYSTEMIC
            bc.font = Font(name="Century Gothic", size=10, bold=True, color="FFFFFF")
            bc.alignment = Alignment(horizontal="left", vertical="center")
            cur_row += 1

        if global_renames:
            ws.row_dimensions[cur_row].height = 22
            desc = "  GLOBAL RENAMES (same value substitution repeated, counted once):  " + "    ·    ".join(
                f"'{old}' → '{new}'  ({cnt}x)"
                for (old, new), cnt in global_renames.items()
            )
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row, end_column=max_col + 3)
            rc = ws.cell(cur_row, 1, desc)
            rc.fill = FILL_RENAME
            rc.font = Font(name="Century Gothic", size=10, bold=True, color="333333")
            rc.alignment = Alignment(horizontal="left", vertical="center")
            cur_row += 1

        # shift all original B row numbers down by however many banner rows we inserted
        ops = [
            (kind, (r_b + banner_rows_needed) if r_b and r_b >= banner_row else r_b, db, da)
            for kind, r_b, db, da in ops
        ]
        insertions = {
            (k + banner_rows_needed if k >= banner_row else k): v
            for k, v in insertions.items()
        }

    # ── Insert deleted rows in-place, bottom→top ─────────────────────────────
    n_removed_rows = sum(len(v) for v in insertions.values()) + len(trailing_dels)
    for b_rn in sorted(insertions.keys(), reverse=True):
        for a_data in reversed(insertions[b_rn]):
            try:
                ws.insert_rows(b_rn)
            except Exception:
                pass
            ws.row_dimensions[b_rn].height = 18
            for c in range(1, max_col + 1):
                val = a_data.get(c, "")
                cell = ws.cell(b_rn, c)
                cell.value = val or None
                cell.fill = FILL_REMOVED
                cell.font = FONT_REMOVED

    # ── Pre-compute row offset (O(log n) lookup) ──────────────────────────────
    sorted_ins = sorted(insertions.items())
    cumulative, total = [], 0
    for b_rn, dl in sorted_ins:
        total += len(dl)
        cumulative.append((b_rn, total))

    def row_offset(r_b):
        lo, hi = 0, len(cumulative)
        while lo < hi:
            mid = (lo + hi) // 2
            if cumulative[mid][0] <= r_b:
                lo = mid + 1
            else:
                hi = mid
        return cumulative[lo - 1][1] if lo > 0 else 0

    # ── Apply highlights with hover comments for old values ──────────────────
    n_changed = n_added_rows = 0
    renames_seen: set = set()   # (old, new) pairs already counted toward n_changed

    for kind, r_b, data_b, data_a in ops:
        if kind == "deleted" or r_b is None:
            continue
        r_adj = r_b + row_offset(r_b)

        if kind == "added":
            n_added_rows += 1
            for c, vb in data_b.items():
                if vb:
                    ws.cell(r_adj, c).fill = FILL_ADDED
                    ws.cell(r_adj, c).font = FONT_ADDED

        elif kind in ("same", "changed"):
            for c in range(1, max_col + 1):
                va = data_a.get(c, "")
                vb = data_b.get(c, "")
                if va != vb:
                    pair = (va, vb)
                    is_global_rename = pair in global_renames
                    if is_global_rename:
                        # Count the whole substitution group as 1 occurrence
                        # toward the headline number, not once per cell.
                        if pair not in renames_seen:
                            renames_seen.add(pair)
                            n_changed += 1
                    else:
                        n_changed += 1
                    cell = ws.cell(r_adj, c)
                    if is_global_rename:
                        cell.fill = FILL_RENAME
                    elif c in systemic_cols:
                        cell.fill = FILL_SYSTEMIC
                    else:
                        cell.fill = FILL_CHANGED
                    cell.font = FONT_CHANGED
                    # Fix #7: systemic/renamed cells still get hover comments
                    try:
                        prev_display = "(formula — value not cached)" if (isinstance(va, str) and va.startswith("=")) else (va or "(empty)")
                        cell.comment = Comment(
                            f"PREVIOUS ({label_a}):\n{prev_display}\n\n— {AUTHOR}",
                            "NJN Comparator", height=60, width=200
                        )
                    except Exception:
                        pass

    # ── Trailing deleted rows ─────────────────────────────────────────────────
    append_at = ws.max_row + 2
    if trailing_dels:
        n_removed_rows = n_removed_rows   # already counted above
        ws.merge_cells(start_row=append_at, start_column=1,
                       end_row=append_at, end_column=max_col + 3)
        hdr = ws.cell(append_at, 1,
                      f"  ROWS REMOVED in {label_b}  (were present in {label_a})")
        hdr.fill = FILL_REMOVED
        hdr.font = Font(name="Century Gothic", size=11, bold=True, color="7B001A")
        hdr.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[append_at].height = 20
        append_at += 1
        for a_data in trailing_dels:
            ws.row_dimensions[append_at].height = 18
            for c in range(1, max_col + 1):
                val = a_data.get(c, "")
                cell = ws.cell(append_at, c, val or None)
                cell.fill = FILL_REMOVED
                cell.font = FONT_REMOVED
            append_at += 1

    systemic_info = [(c, old, new) for c, (old, new) in systemic_cols.items()]
    rename_info = [(old, new, cnt) for (old, new), cnt in global_renames.items()]
    add_legend(ws, ws.max_row + 2, label_a, label_b, max_col, systemic_info, rename_info)
    wb_out.save(out_path)
    return n_changed, n_added_rows, n_removed_rows


def rev_label(path):
    name = os.path.basename(path).replace(".xlsx", "")
    # Sanitize full name for filesystem use — no assumptions about format
    slug = re.sub(r'\s*-\s*', '-', name)   # "x - y" -> "x-y"
    slug = slug.replace(' ', '_')
    slug = re.sub(r'[^\w\-]', '', slug)
    # Full original filename as the display label everywhere
    return slug, name


def main():
    if len(sys.argv) > 1:
        paths = [os.path.expanduser(p) for p in sys.argv[1:]]
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        paths = sorted(glob.glob(os.path.join(script_dir, "*.xlsx")))
        paths = [p for p in paths if "NJN_" not in os.path.basename(p)]

    if not paths:
        print("No .xlsx files found.")
        sys.exit(1)

    files = [(p,) + rev_label(p) for p in paths]
    pairs = [(i, j) for i in range(len(files)) for j in range(i + 1, len(files))]
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir = os.path.dirname(os.path.abspath(__file__))

    for i, j in pairs:
        pa, rev_a, lbl_a = files[i]
        pb, rev_b, lbl_b = files[j]
        out = os.path.join(out_dir, f"NJN_{rev_a}_vs_{rev_b}_{stamp}.xlsx")
        n_chg, n_add, n_del = compare_and_export(pa, pb, out, lbl_a, lbl_b)
        print(f"  {lbl_a} → {lbl_b}: {n_chg} changed, {n_add} added, {n_del} removed")
        print(f"  → {out}")
        if sys.platform == "darwin":
            os.system(f'open "{out}"')


if __name__ == "__main__":
    main()
